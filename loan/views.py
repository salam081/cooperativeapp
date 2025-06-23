from django.shortcuts import render, redirect,get_object_or_404
from django.contrib.auth.decorators import login_required,user_passes_test
from django.template.loader import get_template
from openpyxl.utils import get_column_letter
from django.views.decorators.http import require_http_methods
from django.utils.dateparse import parse_date
from django.template.loader import render_to_string
from datetime import date
from django.db.models import Sum
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse
from xhtml2pdf import pisa
from io import BytesIO
import openpyxl
from django.db.models import Q, Prefetch
from django.core.paginator import Paginator
from django.db.models.functions import ExtractYear
from decimal import Decimal
from django.db import transaction
import pandas as pd
import os
from django.conf import settings
from decimal import Decimal
from collections import defaultdict
from django.db.models.functions import ExtractYear
from .models import LoanRequest, LoanRepayback, LoanType
from django.http import JsonResponse

from .models import *
from accounts.models import *
from main.models import *
from memberapp.models import *
from consumable.models import *




def add_single_loan_payment(request):
    if request.method == "POST":
        ippis = request.POST.get("ippis")
        amount_paid = request.POST.get("amount_paid")
        month = request.POST.get("month")  # Expected format: YYYY-MM-DD

        # Check for missing fields
        if not ippis or not amount_paid or not month:
            messages.error(request, "All fields are required.")
            return redirect(request.path)

        # Validate input
        try:
            ippis = int(ippis)
            amount_paid = Decimal(amount_paid)
            if amount_paid <= 0:
                raise ValueError("Amount must be greater than 0.")
            month = parse_date(month)
            if not month:
                raise ValueError("Invalid date format.")
        except Exception as e:
            messages.error(request, f"Invalid input: {e}")
            return redirect(request.path)

        # Get member
        member = Member.objects.filter(ippis=ippis).first()
        if not member:
            messages.error(request, f"No member found with IPPIS: {ippis}")
            return redirect(request.path)

        # Get approved loan request
        loan_request = LoanRequest.objects.filter(member=member, status='approved').first()
        if not loan_request:
            messages.error(request, "No approved loan request found for this member.")
            return redirect(request.path)

        # Check if payment already made for this month
        already_paid = LoanRepayback.objects.filter(
            loan_request=loan_request,
            repayment_date__year=month.year,
            repayment_date__month=month.month
        ).exists()

        if already_paid:
            messages.warning(request, f"A repayment already exists for {month.strftime('%B %Y')}.")
            return redirect(request.path)

        # Get total already paid
        total_paid = LoanRepayback.objects.filter(loan_request=loan_request).aggregate(
            total=Sum("amount_paid")
        )["total"] or Decimal('0.00')

        remaining_balance = loan_request.approved_amount - total_paid

        if amount_paid > remaining_balance:
            messages.error(request, f"Payment exceeds the remaining balance of ₦{remaining_balance}.")
            return redirect(request.path)

        # Save repayment
        with transaction.atomic():
            new_total_paid = total_paid + amount_paid
            balance_remaining = loan_request.approved_amount - new_total_paid

            LoanRepayback.objects.create(
                loan_request=loan_request,
                amount_paid=amount_paid,
                repayment_date=month,
                balance_remaining=balance_remaining,
                created_by = request.user
            )

            # If fully repaid, update status
            if new_total_paid >= loan_request.approved_amount:
                loan_request.status = 'paid'
                loan_request.save()

        messages.success(request, f"Payment of ₦{amount_paid} recorded successfully for {member}.")
        return redirect(request.path)

    return render(request, "loan/add_single_loan_payment.html")



def get_loan_types_for_year(request):
    year = request.GET.get("year")
    if not year:
        return JsonResponse({"error": "Year not provided"}, status=400)

    loan_types = LoanRequest.objects.filter(
        application_date__year=year
    ).values_list("loan_type__name", flat=True).distinct().order_by("loan_type__name")

    return JsonResponse({"loan_types": list(loan_types)})


def upload_loan_payment(request):
    # Group by year and loan type
    loans = LoanRequest.objects.annotate(year=ExtractYear('application_date')) \
        .values('year', 'loan_type__name').distinct().order_by('-year', 'loan_type__name')
    
    year_to_loan_types = defaultdict(list)
    for loan in loans:
        year_to_loan_types[loan['year']].append(loan['loan_type__name'])

    if request.method == "POST":
        selected_year = request.POST.get("selected_year")
        selected_type = request.POST.get("selected_type")
        file = request.FILES.get("excel_file")

        if not selected_year or not selected_type or not file:
            messages.error(request, "Please select year, loan type, and upload a file.")
            return redirect("upload_loan_payment")

        try:
            selected_year = int(selected_year)
        except ValueError:
            messages.error(request, "Invalid year selected.")
            return redirect("upload_loan_payment")

        try:
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, f"Error reading Excel file: {e}")
            return redirect("upload_loan_payment")

        required_cols = {"IPPIS", "Amount Paid", "Repayment Date"}
        if not required_cols.issubset(df.columns):
            messages.error(request, "Excel must include 'IPPIS', 'Amount Paid', and 'Repayment Date'")
            return redirect("upload_loan_payment")

        # Filter matching loans
        loan_requests = LoanRequest.objects.filter(
            application_date__year=selected_year,
            loan_type__name=selected_type
        ).select_related("member")

        ippis_to_request = {
            str(loan.member.ippis): loan
            for loan in loan_requests
            if loan.member and loan.member.ippis
        }

        uploaded = 0
        skipped = []

        for _, row in df.iterrows():
            ippis = str(row["IPPIS"]).strip()
            amount = row["Amount Paid"]

            try:
                repayment_month = pd.to_datetime(row["Repayment Date"]).date().replace(day=1)
            except Exception:
                skipped.append(ippis)
                continue

            loan_request = ippis_to_request.get(ippis)
            if not loan_request:
                skipped.append(ippis)
                continue

            if LoanRepayback.objects.filter(loan_request=loan_request, repayment_date=repayment_month).exists():
                skipped.append(ippis)
                continue

            # from django.db.models import Sum

            # Calculate total paid before this new repayment
            total_paid_before = LoanRepayback.objects.filter(
                loan_request=loan_request
            ).aggregate(Sum("amount_paid"))["amount_paid__sum"] or Decimal("0.00")

            # Convert current payment to Decimal
            current_payment = Decimal(amount)

            # Calculate new total paid
            new_total_paid = total_paid_before + current_payment

            # Calculate balance
            approved = loan_request.approved_amount or Decimal("0.00")
            # balance_remaining = approved - new_total_paid
            balance_remaining = max(Decimal("0.00"), approved - new_total_paid)

            LoanRepayback.objects.create(
                loan_request=loan_request,
                amount_paid=current_payment,
                repayment_date=repayment_month,
                balance_remaining=balance_remaining,
                created_by=request.user
            )

            uploaded += 1

        messages.success(request, f"{uploaded} loan repayment(s) uploaded.")
        if skipped:
            messages.warning(request, f"Skipped IPPIS: {', '.join(skipped)}")

        return redirect("upload_loan_payment")

    context = {
        "year_to_loan_types": dict(year_to_loan_types),
    }
    return render(request, "loan/upload_loan_payment.html", context)


def filtered_loan_repayments(request):
    years = LoanRequest.objects.annotate(year=ExtractYear("application_date")) \
        .values_list("year", flat=True).distinct().order_by("-year")
    loan_types = LoanRequest.objects.values_list("loan_type__name", flat=True).distinct().order_by("loan_type__name")

    selected_year = request.GET.get("year")
    selected_type = request.GET.get("loan_type")

    filters = Q()
    if selected_year:
        filters &= Q(loan_request__application_date__year=selected_year)
    if selected_type:
        filters &= Q(loan_request__loan_type__name=selected_type)

    repayments_qs = LoanRepayback.objects.select_related("loan_request__member", "loan_request__loan_type") \
        .filter(filters).order_by("-repayment_date")
    # Sum total repayment amount across all filtered records
    total_sum_paid = repayments_qs.aggregate(Sum("amount_paid"))["amount_paid__sum"] or 0

    # Enrich each repayment with total paid and balance
    enriched_repayments = []
    total_sum_remaining = 0 
    for repay in repayments_qs:
        loan = repay.loan_request
        total_paid = LoanRepayback.objects.filter(loan_request=loan).aggregate(Sum("amount_paid"))["amount_paid__sum"] or 0
        approved = loan.approved_amount or 0
        balance = approved - total_paid
        total_sum_remaining += balance  # <-- Add this

        enriched_repayments.append({
            "repayment": repay,
            "total_paid": total_paid,
            "balance_remaining": balance,
        })


    # Add pagination
    paginator = Paginator(enriched_repayments, 5)  # Show 10 per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
    "page_obj": page_obj,
    "years": years,
    "loan_types": loan_types,
    "selected_year": selected_year,
    "selected_type": selected_type,
    "total_sum_paid": total_sum_paid,
    "total_sum_remaining": total_sum_remaining,
    }


    return render(request, "loan/filtered_loan_repayments.html", context)




def get_all_requested_loan(request):
    search_term = request.GET.get('search_term', '').strip()

    # Exclude rejected, approved, and paid
    base_queryset = LoanRequest.objects.exclude(status__in=['rejected', 'approved', 'paid'])

    if search_term:
        results_queryset = base_queryset.filter(
            Q(status__icontains=search_term) |
            Q(member__member__first_name__icontains=search_term) |
            Q(member__member__last_name__icontains=search_term) |
            Q(member__member__username__icontains=search_term) |
            Q(id__icontains=search_term)
        )
    else:
        results_queryset = base_queryset

    results_queryset = results_queryset.order_by('status')

    # Totals by status
    totals_by_status = dict(
        results_queryset.values('status')
        .annotate(total=Sum('approved_amount'))
        .values_list('status', 'total')
    )

    total_approved_amount = results_queryset.aggregate(total=Sum('approved_amount'))['total'] or 0

    totals_by_status = dict(
        results_queryset.values('status')
        .annotate(total=Sum('amount'))
        .values_list('status', 'total')
    )

    total_repaid = LoanRepayback.objects.filter(
        loan_request__in=results_queryset
    ).aggregate(total=Sum('amount_paid'))['total'] or 0

    total_amont_loan_request = totals_by_status.get('approved', 0)
    total_pending = totals_by_status.get('pending', 0)

    paginator = Paginator(results_queryset, 100)
    page_number = request.GET.get('page')
    results = paginator.get_page(page_number)

    if request.GET.get('download_pdf') == '1' and results_queryset.exists():
        if results_queryset.count() > 500:
            return HttpResponse('Too many records to generate PDF. Please narrow your search.', status=400)

        context = {
            'results': results_queryset,
            'search_term': search_term,
            'totals_by_status': totals_by_status,
            'total_approved': total_amont_loan_request,
            'total_pending': total_pending,
            'total_repaid': total_repaid,
            'total_approved_amount': total_approved_amount,
        }
        html = render_to_string('loan/requested_loans_pdf.html', context)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="requested_loans.pdf"'
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('Error generating PDF', status=500)
        return response

    context = {
        'results': results,
        'search_term': "",
        'totals_by_status': totals_by_status,
        'total_approved': total_amont_loan_request,
        'total_pending': total_pending,
        'total_repaid': total_repaid,
        'total_approved_amount': total_approved_amount,
    }

    return render(request, 'loan/get_all_requested_loan.html', context)


def payslip_img_details(request, id):
    payslip_img = LoanRequest.objects.get(id=id)
    context = {'payslip_img': payslip_img}
    return render(request, 'loan/payslip_img_details.html', context)



def edit_requested_loan(request, id):
    loan_types = LoanType.objects.all()
    loanobj = LoanRequest.objects.get(id=id)

    # If user is staff or superuser, use the loan's member
    if request.user.is_staff or request.user.is_superuser:
        member = loanobj.member
    else:
        try:
            member = request.user.member
        except Member.DoesNotExist:
            messages.error(request, "You are not registered as a member.")
            return redirect('some_page')  

        if loanobj.member != member:
            messages.error(request, "You are not allowed to edit this request.")
            return redirect('requested_loan')

    if request.method == "POST":
        loan_type = request.POST['loan_type']
        amount = request.POST['amount']
        loan_term_months = request.POST['loan_term_months']

        LoanRequest.objects.filter(id=id).update(
            member=member,
            loan_type_id=loan_type,
            amount=amount,
            loan_term_months=loan_term_months,
            approved_amount=amount,
        )
        return redirect('requested_loan')

    context = {'loanobj': loanobj, 'loan_types': loan_types}
    return render(request, 'loan/edit_requested_loan.html', context)


def is_admin(user):
    return user.is_staff

@login_required
@user_passes_test(is_admin)
def approve_loan_request(request, id):
    loan_request = get_object_or_404(LoanRequest, id=id, status='pending')
    if request.method == "POST":
        approved_amount = request.POST.get('approved_amount')

        if not approved_amount:
            messages.error(request, "Please enter the approved loan amount.")
            return redirect('approve_loan_request', id=id)

        try:
            approved_amount = float(approved_amount)
            if approved_amount <= 0:
                messages.error(request, "Approved amount must be greater than zero.")
                return redirect('approve_loan_request', id=id)

            if loan_request.loan_type and loan_request.loan_type.max_amount is not None and approved_amount > loan_request.loan_type.max_amount:
                messages.error(request, f"Approved amount cannot exceed the maximum amount for this loan type: {loan_request.loan_type.max_amount}")
                return redirect('approve_loan_request', id=id)

            loan_request.approved_amount = approved_amount
            loan_request.approval_date = timezone.now().date()
            loan_request.status = 'approved'
            loan_request.save()
            messages.success(request, f"Loan request ID {loan_request.id} has been approved for {loan_request.approved_amount}.")
            return redirect('requested_loan') 

        except ValueError:
            messages.error(request, "Invalid approved amount.")
            return redirect('approve_loan_request', id=id)

    context = {'loan_request': loan_request}
    return render(request, 'loan/approve_loan.html', context)



@require_http_methods(["GET", "POST"])
def reject_loan_request(request, id):
    loan_request = LoanRequest.objects.filter(id=id).first()
    if not loan_request:
        messages.error(request, f"No LoanRequest with ID {id} found.")
        return redirect('requested_loan')

    if loan_request.status != 'pending':
        messages.warning(request, f"LoanRequest {id} is already {loan_request.status}. Cannot reject.")
        return redirect('requested_loan')

    if request.method == 'POST':
        reason = request.POST.get('rejection_reason')
        if not reason:
            messages.error(request, "Rejection reason is required.")
            return redirect('reject_loan_request', id=id)

        loan_request.status = 'rejected'
        loan_request.rejection_reason = reason
        loan_request.approval_date = timezone.now().date()
        loan_request.save()

        messages.success(request, f"Loan request ID {loan_request.id} has been rejected with reason.")
        return redirect('requested_loan')

    return render(request, 'loan/reject_loan_form.html', {'loan': loan_request})



def all_reject_loan(request):
    rejected = LoanRequest.objects.filter(status='rejected')
    return render(request,'loan/all_reject_loan.html',{'rejected':rejected} )

def delete_reject_loan(request,id):
    rejectObj = LoanRequest.objects.get(id=id)
    rejectObj.delete()
    return redirect('all_reject_loan')

@login_required
def add_loan_type(request):
    loan_types = LoanType.objects.all()
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        max_amount = request.POST.get('max_amount') or None
        max_loan_term_months = request.POST.get('max_loan_term_months') or None

        if name:
            LoanType.objects.create(
                name=name,
                description=description,
                max_amount=max_amount,
                max_loan_term_months=max_loan_term_months,
                created_by=request.user
            )
            messages.success(request, "Loan type created successfully.")
            return redirect('loan_years_list') 
        else:
            messages.error(request, "Name is required.")
    context = {'loan_types':loan_types}
    return render(request, 'loan/add_loan_type.html',context)

# @login_required
# def delete_loan_type(request, id):
#     loan_type = get_object_or_404(LoanType, id=id)
#     if request.method == 'POST':
#         loan_type.delete()
#         messages.success(request, "Loan type deleted successfully.")
#         return redirect('loan_years_list')  # change to your target view
#     return render(request, 'loan/confirm_delete.html', {'loan_type': loan_type})


def loan_years_list(request):
    # Get distinct year and loan_type combinations
    loans = LoanRequest.objects.annotate(year=ExtractYear('application_date')).values('year', 'loan_type__name').distinct().order_by('-year', 'loan_type__name')

    # Structure the data as {2025: ['LONG TERM LOAN'], 2024: ['SHORT TERM LOAN', ...]}
    year_to_loan_types = {}
    for loan in loans:
        year = loan['year']
        loan_type = loan['loan_type__name']
        year_to_loan_types.setdefault(year, []).append(loan_type)

    context = {'year_to_loan_types': year_to_loan_types,}
    return render(request, "loan/loan_years_list.html", context)


def loans_by_year(request, year, loan_type_filter):
    loan_type = get_object_or_404(LoanType, name__iexact=loan_type_filter)
    status_filter = request.GET.get('status')

    # Filter loans by type and year
    loanobj = LoanRequest.objects.filter(loan_type=loan_type, date_created__year=year)

    # Optional: Filter by status if given
    if status_filter:
        loanobj = loanobj.filter(status__iexact=status_filter)

    # Totals by status
    totals_by_status = dict(
        loanobj.values('status')
        .annotate(total=Sum('approved_amount'))
        .values_list('status', 'total')
    )

    context = {
        'year': year,
        'loan_type': loan_type,
        'loanobj': loanobj,
        'totals_by_status': totals_by_status,
        'selected_status': status_filter, 
    }

    # Handle PDF download
    if request.GET.get('download') == 'pdf':
        template_path = 'loan/loans_by_year_pdf.html'
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="loans_{loan_type.name}_{year}.pdf"'
        template = get_template(template_path)
        html = template.render(context)
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('We had some errors <pre>' + html + '</pre>')
        return response

    # Handle Excel download
    if request.GET.get('download') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Loan Data"

        headers = ['ID', 'Applicant', 'Amount', 'Approved Amount', 'Account Number', 'Bank Name', 'Bank Code', 'Status', 'Date Created']
        ws.append(headers)

        for loan in loanobj:
            ws.append([
                loan.id,
                str(loan.member),
                loan.amount,
                loan.approved_amount,
                loan.account_number,
                str(loan.bank_name),
                str(loan.bank_code),
                loan.status,
                loan.date_created.strftime('%Y-%m-%d')
            ])

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="loans_{loan_type.name}_{year}.xlsx"'
        wb.save(response)
        return response

    # Render normal HTML page
    return render(request, "loan/loans_by_year.html", context)
